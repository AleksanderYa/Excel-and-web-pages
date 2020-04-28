
import webbrowser


def position():
    for i in range(ou,500):
        yield i
position = position()


def wr(ryad, col, value):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filename='./Food.xlsx' )
        ws = wb.worksheets[0]
        print()
        answer = int(input('if yes-1 or 0 if not?\n'))
        if answer == 0:
           # print('_zero')
            ws.cell(ryad , col).value = 'no'
            # print('zero')
        elif answer == 1:
            # print('_one')
            answer2 = str(input('Coments:\n'))
            ws.cell(ryad , 5).value = answer2
            ws.cell(ryad , col).value = 'yes'
    except Exception as e:
        print(e)
    finally:
        wb.save('./Food.xlsx')


def comp():
    try:   
        from xlrd import open_workbook
        book = open_workbook('./Food.xlsx')
        sheet = book.sheet_by_index(0) 
        for i in range(ou,500):
            if sheet.row_values(i)[1]: 
                a = sheet.row_values(i)[1]
                url = 'http://' + a[4::]
                webbrowser.open(url, autoraise=True)
                print()
                # wr = wr(next(position2),2, url, 2)
                yield print(url),wr(next(position), 3, url)
                
            else:
                continue
    except Exception as e:
        print(e)
comp = comp()

ou = int(input('С какой позиции начнем?\n'))
for i in range(1, 500):
    answer = str(input('put Enter to continue, or \'q\' (\'Q\', \'exit\') for EXIT\n'))
    if not answer:
        next(comp)
    elif answer == 'q' or 'Q'or 'exit':
        exit()
